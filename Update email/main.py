'''
Name: Kumo Pascaline Myenneh
Email: kumo.pascaline@ictuniversity.edu.cm
ICTU Matricule: ICTU20201168
Contact: 661253212
Course: Programming in python
Course Instructor: Mr Fru Emmanuel
this is a python  program that access, edits and update emails in an excel file from
 @helpinghands.org to @helpinghands.cm
'''

import openpyxl as xl

wb = xl.load_workbook('employees_data.xlsx')
sheet = wb['Sheet1'] # sheet 1 is the active cell thats y it is called in the workbook(wb)


old_domain = 'helpinghands.cm '
new_domain = 'handsinhands.org'

# ------- CREATING THE EMAILS USING THE USER'S NAME------------#


for row in range (2, sheet.max_row + 1):# begins from two because we're working with the second row elements

    cell = sheet.cell(row, 2 )# reads the element of column 1 in all the rows

    new_email = ((cell.value).lower() + old_domain ) # reads the element of the cell e.g jude and add @helpinghands.cm to it
    #----- creating a new cell to store the new email -----#

    new_email_cell = sheet.cell(row, 4) # stores the new email in the 4th column(column D)

    #------ attributing the data to the new cells (column) created -----#
    new_email_cell.value = new_email
    
# ----- creating a new database with the .csv extention that stores the new and old email 
wb.save('employees_data.xlsx')

# ----- creating a new row for updated email -----#


for i in range(2,sheet.max_row + 1):#Begins from 2 because we are not working with the first row which is the header.
    cell = sheet.cell(i,4)#This tells the compiler to access the elements of the 3rd column
    if old_domain in cell.value:
        updated_email=(cell.value).replace(old_domain,new_domain)#replacing the old emails with the new emails

        sheet.cell(i,4).value = updated_email
wb.save('updated_email.csv')#creating the a csv file
